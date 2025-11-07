import pytest
from openpyxl import load_workbook

def read_excel_data(file_path):
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active
    data = []

    # Lewati baris header pertama
    for row in sheet.iter_rows(min_row=2, values_only=True):
        username, password, expected_text = row
        data.append((username or "", password or "", expected_text))
    return data


@pytest.mark.parametrize(
    "username,password,expected_text",
    read_excel_data("tests/data/login_data.xlsx")
)
def test_login_excel(page, username, password, expected_text):
    print(f"\nUji username={username}, password={password}")

    # Simulasi login hasil
    if username == "admin" and password == "12345":
        result = "Login successful"
    elif username == "":
        result = "Username required"
    elif password == "":
        result = "Password required"
    else:
        result = "Invalid password"

    assert result == expected_text, f"Dapat: {result}, harapan: {expected_text}"
